import openpyxl
import openpyxl.utils
#python3 -m pip install openpyxl | pip3 install openpyxl

wb = openpyxl.load_workbook(r"C:/Users/g-oli/Desktop/Projekt ZF/Instruktionen/Verfahrensbeschreibung.xlsx")
sheet = wb.active
test = sheet.calculate_dimension()
num_rows = sheet.max_row - 1
num_cols = sheet.max_column
new_char = ""
new_charB = ""
new_charC = ""
new_charD = ""
new_charE = ""
new_charF = ""
new_charG = ""

for y in range(1, num_rows):
    for x in range (1, num_cols):
        char = openpyxl.utils.get_column_letter(y)
        char = char + str(x)
        if sheet[char].value != None:
            if char[1] == "1":
                new_char += (char + " " + str(sheet[char].value + "\t | \t"))
            if char[1] == "2":
                new_charB += (char + " " + str(sheet[char].value + "\t | \t"))
            if char[1] == "3":
                new_charC += (char + " " + str(sheet[char].value + "\t | \t"))
            if char[1] == "4":
                new_charD += (char + " " + str(sheet[char].value + "\t | \t"))
            if char[1] == "5":
                new_charE += (char + " " + str(sheet[char].value + "\t | \t"))
            if char[1] == "6":
                new_charF += (char + " " + str(sheet[char].value + "\t | \t"))
            if char[1] == "7":
                new_charG += (char + " " + str(sheet[char].value + "\t | \t"))
print(new_char)
print(new_charB)
print(new_charC)
print(new_charD)
print(new_charE)
print(new_charF)
print(new_charG)

